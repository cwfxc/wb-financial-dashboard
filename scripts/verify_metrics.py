"""
Полная независимая проверка всех метрик дашборда.
Подключается к тем же источникам (Postgres + Excel),
считает всё на Python и сравнивает с Power BI.
"""
import psycopg2
from openpyxl import load_workbook
from collections import defaultdict

# ========== 1. ПОДКЛЮЧЕНИЕ К ИСТОЧНИКАМ ==========

# PostgreSQL
conn = psycopg2.connect(
    host="109.73.206.144",
    port=5433,
    database="u115241_test_ba",
    user="u115241_powerbi",
    password="QKgbytMrmtA8T38m"
)
cur = conn.cursor()

# Excel (себестоимость)
wb = load_workbook("file2.xlsx", data_only=True)

# ========== 2. ЗАГРУЗКА СЕБЕСТОИМОСТИ ==========
ws_cost = wb["Себестоимость"]
cost_rows = list(ws_cost.iter_rows(values_only=True))
costs = {}  # barcode -> cost
for row in cost_rows[1:]:
    if row[0] is not None and row[1] is not None:
        bc = str(int(row[0]))
        costs[bc] = float(row[1])

print(f"=== Себестоимость ===")
print(f"Загружено баркодов с себестоимостью: {len(costs)}")

# ========== 3. СПРАВОЧНИК НОМЕНКЛАТУРЫ ==========
ws_prod = wb["Справочник номенклатуры"]
prod_rows = list(ws_prod.iter_rows(values_only=True))
prod_header = prod_rows[0]
barcode_idx = list(prod_header).index("barcode")
article_idx = list(prod_header).index("supplierArticle")
subject_idx = list(prod_header).index("subject")

products = {}  # barcode -> {article, subject}
for row in prod_rows[1:]:
    if row[article_idx] is not None and row[barcode_idx] is not None:
        bc = str(int(row[barcode_idx]))
        products[bc] = {
            'article': row[article_idx],
            'subject': row[subject_idx]
        }

print(f"Загружено товаров: {len(products)}")

# ========== 4. ПРОВЕРКА FactSalesReport ==========
print("\n=== FactSalesReport (v_report_realization) ===")

# Общее количество строк
cur.execute("SELECT COUNT(*) FROM v_report_realization")
total_rows = cur.fetchone()[0]
print(f"Всего строк: {total_rows:,}")

# Строки по типам
cur.execute("""
    SELECT doc_type_name, COUNT(*), SUM(quantity), SUM(retail_amount)
    FROM v_report_realization
    GROUP BY doc_type_name
    ORDER BY COUNT(*) DESC
""")
print("\nРазбивка по doc_type_name:")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]:,} строк, quantity={row[2]:,.0f}, retail_amount={row[3]:,.2f}")

# ========== 5. ВЫРУЧКА WB ==========
cur.execute("""
    SELECT
        COALESCE(SUM(CASE WHEN doc_type_name = 'Продажа' THEN retail_amount ELSE 0 END), 0) -
        COALESCE(SUM(CASE WHEN doc_type_name = 'Возврат' THEN retail_amount ELSE 0 END), 0)
    FROM v_report_realization
""")
revenue = cur.fetchone()[0]
print(f"\n=== МЕТРИКИ ===")
print(f"Выручка WB: {float(revenue):,.2f}")

# ========== 6. ЗАКАЗОВ ШТ ==========
cur.execute("SELECT SUM(quantity) FROM v_orders")
orders_qty = cur.fetchone()[0]
print(f"Заказов шт: {float(orders_qty):,.0f}")

# ========== 7. КОМИССИЯ WB ==========
cur.execute("SELECT SUM(ppvz_sales_commission) FROM v_report_realization")
commission = cur.fetchone()[0]
print(f"Комиссия WB: {float(commission):,.2f}")

# ========== 8. ЛОГИСТИКА ==========
cur.execute("SELECT SUM(delivery_rub) FROM v_report_realization")
logistics = cur.fetchone()[0]
print(f"Логистика: {float(logistics):,.2f}")

# ========== 9. СЕБЕСТОИМОСТЬ (ключевая проверка) ==========
# Получаем все строки Продажа с баркодами
cur.execute("""
    SELECT barcode, quantity, retail_amount, ppvz_reward, ppvz_vw, ppvz_vw_nds, delivery_rub
    FROM v_report_realization
    WHERE doc_type_name = 'Продажа'
""")
sale_rows = cur.fetchall()
print(f"\nСтрок 'Продажа': {len(sale_rows):,}")

total_cost = 0
total_qty_with_cost = 0
total_qty_no_cost = 0
missing_barcodes = set()
cost_breakdown = defaultdict(lambda: {'qty': 0, 'cost_total': 0, 'revenue': 0})

for row in sale_rows:
    bc = str(row[0]) if row[0] else None
    qty = float(row[1]) if row[1] else 0
    retail = float(row[2]) if row[2] else 0

    if bc and bc in costs and qty != 0:
        item_cost = costs[bc] * qty
        total_cost += item_cost
        total_qty_with_cost += qty

        article = products.get(bc, {}).get('article', 'UNKNOWN')
        cost_breakdown[article]['qty'] += qty
        cost_breakdown[article]['cost_total'] += item_cost
        cost_breakdown[article]['revenue'] += retail
    elif bc and bc not in costs and qty != 0:
        total_qty_no_cost += qty
        missing_barcodes.add(bc)

print(f"Расход себестоимость (Python): {total_cost:,.2f}")
print(f"Кол-во с себестоимостью: {total_qty_with_cost:,.0f}")
print(f"Кол-во без себестоимости: {total_qty_no_cost:,.0f}")
print(f"Баркодов без себестоимости: {len(missing_barcodes)}")

# ========== 10. ЭКВАЙРИНГ ==========
total_acquiring = 0
for row in sale_rows:
    qty = float(row[1]) if row[1] else 0
    retail = float(row[2]) if row[2] else 0
    if qty != 0:
        total_acquiring += (retail / qty) * 0.02
    # qty=0 => DIVIDE(retail, 0, 0) * 0.02 = 0 (как в DAX)

print(f"Расход эквайринг (Python): {total_acquiring:,.2f}")

# ========== 11. ВАЛОВАЯ ПРИБЫЛЬ ==========
# Продажа: вычитаем расходы
cur.execute("""
    SELECT barcode, quantity, retail_amount, ppvz_reward, ppvz_vw, ppvz_vw_nds, delivery_rub, ppvz_sales_commission
    FROM v_report_realization
    WHERE doc_type_name = 'Продажа'
""")
sale_rows_full = cur.fetchall()

gross_expenses_sale = 0
for row in sale_rows_full:
    bc = str(row[0]) if row[0] else None
    qty = float(row[1]) if row[1] else 0
    retail = float(row[2]) if row[2] else 0
    reward = float(row[3]) if row[3] else 0
    vw = float(row[4]) if row[4] else 0
    vw_nds = float(row[5]) if row[5] else 0
    delivery = float(row[6]) if row[6] else 0

    # Себестоимость
    cost_val = 0
    if bc and bc in costs:
        cost_val = costs[bc] * qty

    # Эквайринг
    acq = 0
    if qty != 0:
        acq = (retail / qty) * 0.02

    gross_expenses_sale += reward + vw + vw_nds + delivery + acq + cost_val

# Возврат: прибавляем обратно
cur.execute("""
    SELECT barcode, quantity, retail_amount, ppvz_reward, ppvz_vw, ppvz_vw_nds, delivery_rub
    FROM v_report_realization
    WHERE doc_type_name = 'Возврат'
""")
return_rows = cur.fetchall()

gross_expenses_return = 0
for row in return_rows:
    bc = str(row[0]) if row[0] else None
    qty = float(row[1]) if row[1] else 0
    retail = float(row[2]) if row[2] else 0
    reward = float(row[3]) if row[3] else 0
    vw = float(row[4]) if row[4] else 0
    vw_nds = float(row[5]) if row[5] else 0
    delivery = float(row[6]) if row[6] else 0

    cost_val = 0
    if bc and bc in costs:
        cost_val = costs[bc] * qty

    acq = 0
    if qty != 0:
        acq = (retail / qty) * 0.02

    gross_expenses_return += reward + vw + vw_nds + delivery + acq + cost_val

gross_profit = float(revenue) - gross_expenses_sale + gross_expenses_return
print(f"\n=== ВАЛОВАЯ ПРИБЫЛЬ ===")
print(f"Выручка: {float(revenue):,.2f}")
print(f"Расходы Продажа: {gross_expenses_sale:,.2f}")
print(f"Расходы Возврат (возвращаются): {gross_expenses_return:,.2f}")
print(f"Валовая прибыль (Python): {gross_profit:,.2f}")

# ========== 12. ТОП УБЫТОЧНЫХ ТОВАРОВ ==========
print(f"\n=== ТОП-10 убыточных артикулов (по себестоимости vs выручка) ===")
sorted_articles = sorted(cost_breakdown.items(), key=lambda x: x[1]['revenue'] - x[1]['cost_total'])
for article, data in sorted_articles[:10]:
    profit = data['revenue'] - data['cost_total']
    print(f"  {article}: qty={data['qty']:,.0f}, выручка={data['revenue']:,.0f}, себест={data['cost_total']:,.0f}, разница={profit:,.0f}")

# ========== 13. ПРОВЕРКА КОЛИЧЕСТВА СТРОК С quantity=0 ==========
cur.execute("""
    SELECT COUNT(*) FROM v_report_realization
    WHERE doc_type_name = 'Продажа' AND (quantity = 0 OR quantity IS NULL)
""")
zero_qty = cur.fetchone()[0]
print(f"\n=== Строки Продажа с quantity=0: {zero_qty:,} ===")

cur.execute("""
    SELECT COUNT(*) FROM v_report_realization
    WHERE doc_type_name = 'Продажа' AND quantity > 0
""")
pos_qty = cur.fetchone()[0]
print(f"Строки Продажа с quantity > 0: {pos_qty:,}")

cur.execute("""
    SELECT COUNT(*) FROM v_report_realization
    WHERE doc_type_name = 'Продажа' AND quantity < 0
""")
neg_qty = cur.fetchone()[0]
print(f"Строки Продажа с quantity < 0: {neg_qty:,}")

# ========== 14. СРЕДНИЕ ЦЕНЫ ==========
print(f"\n=== Средние цены ===")
avg_cost = total_cost / total_qty_with_cost if total_qty_with_cost else 0
avg_price = float(revenue) / total_qty_with_cost if total_qty_with_cost else 0
print(f"Средняя себестоимость за шт: {avg_cost:,.2f}")
print(f"Средняя цена продажи за шт: {avg_price:,.2f}")
print(f"Соотношение себестоимость/цена: {avg_cost/avg_price:.2f}x")

# ========== 15. ПРОВЕРКА ДУБЛИКАТОВ ==========
print(f"\n=== Проверка дубликатов ===")
cur.execute("""
    SELECT COUNT(*) as total, COUNT(DISTINCT (date::text || barcode::text || quantity::text || retail_amount::text)) as distinct_combos
    FROM v_report_realization
    WHERE doc_type_name = 'Продажа'
""")
dup_check = cur.fetchone()
print(f"Всего строк Продажа: {dup_check[0]:,}")
print(f"Уникальных комбинаций (дата+баркод+кол+сумма): {dup_check[1]:,}")

conn.close()
print("\n✓ Проверка завершена")
