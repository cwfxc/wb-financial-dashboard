"""Complete verification: all metrics for full period and March 2022."""
import psycopg2
from openpyxl import load_workbook

conn = psycopg2.connect(
    host="109.73.206.144", port=5433,
    database="u115241_test_ba",
    user="u115241_powerbi", password="QKgbytMrmtA8T38m"
)
cur = conn.cursor()

# Load costs
wb = load_workbook("file2.xlsx", data_only=True)
ws = wb["Себестоимость"]
costs = {}
for row in list(ws.iter_rows(values_only=True))[1:]:
    if row[0] is not None and row[1] is not None:
        costs[str(int(row[0]))] = float(row[1])

def calc_metrics(date_filter=""):
    """Calculate all metrics with optional date filter."""
    where = f"WHERE sale_dt >= '{date_filter[0]}' AND sale_dt <= '{date_filter[1]}'" if date_filter else ""

    # 1. Vyruchka WB
    cur.execute(f"""
        SELECT
            COALESCE(SUM(CASE WHEN doc_type_name = 'Продажа' THEN retail_amount ELSE 0 END), 0) -
            COALESCE(SUM(CASE WHEN doc_type_name = 'Возврат' THEN retail_amount ELSE 0 END), 0)
        FROM "SalesReport" {where}
    """)
    revenue = float(cur.fetchone()[0])

    # 2. Komissiya WB
    cur.execute(f'SELECT COALESCE(SUM(ppvz_sales_commission), 0) FROM "SalesReport" {where}')
    commission = float(cur.fetchone()[0])

    # 3. Logistika
    cur.execute(f'SELECT COALESCE(SUM(delivery_rub), 0) FROM "SalesReport" {where}')
    logistics = float(cur.fetchone()[0])

    # 4. Vozvraty
    cur.execute(f"""
        SELECT COALESCE(SUM(quantity), 0) FROM "SalesReport"
        {where + ' AND' if where else 'WHERE'} doc_type_name = 'Возврат'
    """)
    returns = float(cur.fetchone()[0])

    # 5. Valovaya pribyl
    sale_where = where + " AND" if where else "WHERE"
    cur.execute(f"""
        SELECT barcode, quantity, retail_amount, ppvz_reward, ppvz_vw, ppvz_vw_nds, delivery_rub
        FROM "SalesReport" {sale_where} doc_type_name = 'Продажа'
    """)
    exp_sale = 0
    cost_sale = 0
    for row in cur.fetchall():
        bc = str(row[0]) if row[0] else None
        qty = float(row[1]) if row[1] else 0
        retail = float(row[2]) if row[2] else 0
        reward = float(row[3]) if row[3] else 0
        vw = float(row[4]) if row[4] else 0
        vw_nds = float(row[5]) if row[5] else 0
        delivery = float(row[6]) if row[6] else 0
        c = costs.get(bc, 0) * qty if bc else 0
        acq = (retail / qty) * 0.02 if qty != 0 else 0
        exp_sale += reward + vw + vw_nds + delivery + acq + c
        cost_sale += c

    cur.execute(f"""
        SELECT barcode, quantity, retail_amount, ppvz_reward, ppvz_vw, ppvz_vw_nds, delivery_rub
        FROM "SalesReport" {sale_where} doc_type_name = 'Возврат'
    """)
    exp_ret = 0
    cost_ret = 0
    for row in cur.fetchall():
        bc = str(row[0]) if row[0] else None
        qty = float(row[1]) if row[1] else 0
        retail = float(row[2]) if row[2] else 0
        reward = float(row[3]) if row[3] else 0
        vw = float(row[4]) if row[4] else 0
        vw_nds = float(row[5]) if row[5] else 0
        delivery = float(row[6]) if row[6] else 0
        c = costs.get(bc, 0) * qty if bc else 0
        acq = (retail / qty) * 0.02 if qty != 0 else 0
        exp_ret += reward + vw + vw_nds + delivery + acq + c
        cost_ret += c

    gross = revenue - exp_sale + exp_ret
    rentab = gross / revenue * 100 if revenue != 0 else 0
    roi = gross / (cost_sale - cost_ret) * 100 if (cost_sale - cost_ret) != 0 else 0

    return {
        'revenue': revenue,
        'commission': commission,
        'logistics': logistics,
        'returns': returns,
        'gross_profit': gross,
        'rentab': rentab,
        'roi': roi,
        'cost_sale': cost_sale,
        'cost_ret': cost_ret
    }

# === FULL PERIOD ===
print("=" * 60)
print("FULL PERIOD (all dates)")
print("=" * 60)
m = calc_metrics()
print(f"Vyruchka WB:     {m['revenue']:>15,.2f}")
print(f"Komissiya WB:    {m['commission']:>15,.2f}")
print(f"Logistika:       {m['logistics']:>15,.2f}")
print(f"Valovaya pribyl: {m['gross_profit']:>15,.2f}")
print(f"Rentabelnost:    {m['rentab']:>10,.2f}%")
print(f"ROI:             {m['roi']:>10,.2f}%")
print(f"Vozvraty sht:    {m['returns']:>10,.0f}")

# === MARCH 2022 ===
print(f"\n{'=' * 60}")
print("MARCH 2022 (01.03 - 31.03)")
print("=" * 60)
m2 = calc_metrics(("2022-03-01", "2022-03-31"))
print(f"Vyruchka WB:     {m2['revenue']:>15,.2f}")
print(f"Komissiya WB:    {m2['commission']:>15,.2f}")
print(f"Logistika:       {m2['logistics']:>15,.2f}")
print(f"Valovaya pribyl: {m2['gross_profit']:>15,.2f}")
print(f"Rentabelnost:    {m2['rentab']:>10,.2f}%")
print(f"Vozvraty sht:    {m2['returns']:>10,.0f}")

# === ZAKAZOV SHT (from v_orders) ===
print(f"\n{'=' * 60}")
print("ZAKAZOV SHT CHECK")
print("=" * 60)
cur.execute("SELECT SUM(quantity) FROM v_orders")
total_orders = float(cur.fetchone()[0])
print(f"Full period: {total_orders:,.0f}")

cur.execute("SELECT SUM(quantity) FROM v_orders WHERE date >= '2022-03-01' AND date <= '2022-03-31'")
march_orders = cur.fetchone()[0]
print(f"March 2022: {float(march_orders) if march_orders else 0:,.0f}")

# === KOMISSIYA % (from Sales) ===
print(f"\n{'=' * 60}")
print("KOMISSIYA WB % CHECK")
print("=" * 60)
cur.execute('SELECT SUM("priceWithDisc") FROM "Sales" WHERE "IsStorno" = 0')
full_pd = float(cur.fetchone()[0])
print(f"Full period: {m['commission']:,.2f} / {full_pd:,.2f} = {m['commission']/full_pd*100:.2f}%")

cur.execute("""
    SELECT SUM("priceWithDisc") FROM "Sales"
    WHERE "IsStorno" = 0 AND date >= '2022-03-01' AND date < '2022-04-01'
""")
mar_pd = float(cur.fetchone()[0])
print(f"March: {m2['commission']:,.2f} / {mar_pd:,.2f} = {m2['commission']/mar_pd*100:.2f}%")

# === PRODAZHI PO OBLASTYAM (March) ===
print(f"\n{'=' * 60}")
print("PRODAZHI PO OBLASTYAM - March 2022")
print("=" * 60)
cur.execute("""
    SELECT "oblastOkrugName", SUM("forPay")
    FROM "Sales"
    WHERE "isRealization" = true AND date >= '2022-03-01' AND date < '2022-04-01'
    GROUP BY "oblastOkrugName"
    ORDER BY SUM("forPay") DESC
    LIMIT 6
""")
for row in cur.fetchall():
    print(f"  {row[0]}: {float(row[1]):,.0f}")

# === TOP DECLINE LEADERS - March 2022 ===
print(f"\n{'=' * 60}")
print("TOP DECLINE LEADERS - March 2022 vs previous 31 days")
print("=" * 60)

# Load product reference for article names
ws_prod = wb["Справочник номенклатуры"]
prod_rows = list(ws_prod.iter_rows(values_only=True))
header = prod_rows[0]
bc_idx = list(header).index("barcode")
art_idx = list(header).index("supplierArticle")
products = {}
for row in prod_rows[1:]:
    if row[art_idx] and row[bc_idx]:
        products[str(int(row[bc_idx]))] = row[art_idx]

# March = days 1-31, prev = Jan 29 - Feb 28 (31 days before March 1)
cur.execute("""
    SELECT barcode,
        SUM(CASE WHEN sale_dt >= '2022-01-29' AND sale_dt <= '2022-02-28'
                 AND doc_type_name = 'Продажа' THEN retail_amount ELSE 0 END) -
        SUM(CASE WHEN sale_dt >= '2022-01-29' AND sale_dt <= '2022-02-28'
                 AND doc_type_name = 'Возврат' THEN retail_amount ELSE 0 END) as prev_rev,
        SUM(CASE WHEN sale_dt >= '2022-03-01' AND sale_dt <= '2022-03-31'
                 AND doc_type_name = 'Продажа' THEN retail_amount ELSE 0 END) -
        SUM(CASE WHEN sale_dt >= '2022-03-01' AND sale_dt <= '2022-03-31'
                 AND doc_type_name = 'Возврат' THEN retail_amount ELSE 0 END) as cur_rev
    FROM "SalesReport"
    WHERE sale_dt >= '2022-01-29' AND sale_dt <= '2022-03-31'
    GROUP BY barcode
""")

article_data = {}
for row in cur.fetchall():
    bc = str(row[0]) if row[0] else None
    if bc and bc in products:
        art = products[bc]
        if art not in article_data:
            article_data[art] = {'prev': 0, 'cur': 0}
        article_data[art]['prev'] += float(row[1] or 0)
        article_data[art]['cur'] += float(row[2] or 0)

deltas = [(art, d['prev'], d['cur'], d['cur'] - d['prev']) for art, d in article_data.items()]
deltas.sort(key=lambda x: x[3])

print("\nTOP-5 DECLINE:")
for art, prev, cur, delta in deltas[:5]:
    a = art.replace("supplierArticle ", "Art. ")
    print(f"  {a}: prev={prev:,.0f}, cur={cur:,.0f}, delta={delta:,.0f}")

print("\nTOP-5 GROWTH:")
for art, prev, cur, delta in deltas[-5:]:
    a = art.replace("supplierArticle ", "Art. ")
    print(f"  {a}: prev={prev:,.0f}, cur={cur:,.0f}, delta={delta:,.0f}")

# === MONTHLY SUMMARY ===
print(f"\n{'=' * 60}")
print("MONTHLY REVENUE SUMMARY (sale_dt)")
print("=" * 60)
cur.execute("""
    SELECT date_trunc('month', sale_dt)::date,
        SUM(CASE WHEN doc_type_name = 'Продажа' THEN retail_amount ELSE 0 END) -
        SUM(CASE WHEN doc_type_name = 'Возврат' THEN retail_amount ELSE 0 END)
    FROM "SalesReport"
    WHERE sale_dt >= '2021-12-01'
    GROUP BY date_trunc('month', sale_dt)
    ORDER BY 1
""")
for row in cur.fetchall():
    print(f"  {row[0]}: {float(row[1]):>12,.0f}")

# === RELATIONSHIP CHAIN CHECK ===
print(f"\n{'=' * 60}")
print("RELATIONSHIP CHAIN: barcode -> product -> cost")
print("=" * 60)
# Count barcodes in SalesReport that have matching cost
cur.execute("""
    SELECT COUNT(DISTINCT barcode) FROM "SalesReport"
    WHERE doc_type_name = 'Продажа' AND quantity > 0
""")
sr_barcodes = cur.fetchone()[0]

matched = sum(1 for bc in set() or [])  # placeholder

cur.execute("""
    SELECT COUNT(DISTINCT barcode) FROM "SalesReport"
    WHERE doc_type_name = 'Продажа' AND quantity > 0
""")
total_bc = cur.fetchone()[0]

cur.execute("""
    SELECT DISTINCT barcode FROM "SalesReport"
    WHERE doc_type_name = 'Продажа' AND quantity > 0
""")
sr_bcs = set(str(r[0]) for r in cur.fetchall() if r[0])
matched_cost = sr_bcs & set(costs.keys())
matched_prod = sr_bcs & set(products.keys())
no_cost = sr_bcs - set(costs.keys())
no_prod = sr_bcs - set(products.keys())

print(f"Unique barcodes in SalesReport (Prodazha, qty>0): {len(sr_bcs)}")
print(f"Matched in DimCost: {len(matched_cost)} ({len(matched_cost)/len(sr_bcs)*100:.1f}%)")
print(f"Matched in DimProduct: {len(matched_prod)} ({len(matched_prod)/len(sr_bcs)*100:.1f}%)")
print(f"Missing cost: {len(no_cost)} barcodes")
print(f"Missing product: {len(no_prod)} barcodes")
if no_cost:
    print(f"  Barcodes without cost: {list(no_cost)[:5]}")

conn.close()
print("\nDone!")
